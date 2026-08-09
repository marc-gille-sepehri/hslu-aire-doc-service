"""Regenerate tests/fixtures/*.xlsx.

Run manually after changing a fixture; the .xlsx files are committed as binaries.

openpyxl never evaluates formulas, so a workbook it writes carries no cached results
and load_workbook(data_only=True) returns None for every formula cell. The acceptance
fixture needs the computed values (§6), so the cached <v> elements are injected into
the sheet XML afterwards — the same thing Excel would have written on save.
"""
import io
import pathlib
import re
import shutil
import zipfile

from openpyxl import Workbook

FIXTURES = pathlib.Path(__file__).parent / "fixtures"


def inject_cached(raw: bytes, part: str, values: dict[str, object]) -> bytes:
    """Give each formula cell in `values` the cached result Excel would have stored."""
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        items = {n: z.read(n) for n in z.namelist()}

    xml = items[part].decode("utf-8")
    for coord, value in values.items():
        # openpyxl emits <c r="B7"><f>SUM(B4:B5)</f><v></v></c> — an empty cached slot.
        # Fill it (or add one) so data_only=True sees a result, as Excel would write.
        pattern = re.compile(
            rf'(<c r="{coord}"[^>]*>\s*<f[^>]*>.*?</f>)(\s*<v\s*/>|\s*<v>.*?</v>)?(\s*</c>)', re.S
        )
        xml, n = pattern.subn(rf"\g<1><v>{value}</v>\g<3>", xml)
        if n != 1:
            raise SystemExit(f"could not inject cached value for {coord} (matched {n})")
    items[part] = xml.encode("utf-8")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for name, blob in items.items():
            z.writestr(name, blob)
    return out.getvalue()


def objektliste() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"

    ws["A1"] = "Objektliste 2026"
    ws.merge_cells("A1:D1")
    ws["B2"] = "Fläche"
    ws.merge_cells("B2:C2")
    ws["D2"] = "Ertrag"
    for col, head in zip("ABCD", ["Liegenschaft", "Wohnen", "Gewerbe", "Miete/Jahr"]):
        ws[f"{col}3"] = head
    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Weidgasse 14", 820, 140, 196400
    ws["A5"], ws["B5"], ws["D5"] = "Weidgasse 16", 760, 172800
    # row 6 stays entirely empty on purpose — the gap is the assertion
    ws["A7"] = "Zwischensumme"
    ws["B7"], ws["C7"], ws["D7"] = "=SUM(B4:B5)", "=SUM(C4:C5)", "=SUM(D4:D5)"

    buf = io.BytesIO()
    wb.save(buf)
    return inject_cached(
        buf.getvalue(),
        "xl/worksheets/sheet1.xml",
        {"B7": 1580, "C7": 140, "D7": 369200},
    )


if __name__ == "__main__":
    if FIXTURES.exists():
        shutil.rmtree(FIXTURES)
    FIXTURES.mkdir(parents=True)
    (FIXTURES / "objektliste.xlsx").write_bytes(objektliste())
    print(f"wrote {FIXTURES / 'objektliste.xlsx'}")
