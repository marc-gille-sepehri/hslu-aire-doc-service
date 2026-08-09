"""Acceptance tests for the cell-addressed format — docs/delta-spec-doc-convert-cells.md §6.

Fixtures live in tests/fixtures/ as committed binaries; regenerate with
    python3 tests/make_fixtures.py
"""
import io
import pathlib
import re

import pytest
from openpyxl import Workbook

from app.cells import build_cells, serialize_csv, serialize_xlsx

FIXTURES = pathlib.Path(__file__).parent / "fixtures"
OBJEKTLISTE = (FIXTURES / "objektliste.xlsx").read_bytes()

EXPECTED_SILENT = """# Sheet: Q3
R1| A[merged A1:D1]:Objektliste 2026
R2| B[merged B2:C2]:Fläche || D:Ertrag
R3| A:Liegenschaft || B:Wohnen || C:Gewerbe || D:Miete/Jahr
R4| A:Weidgasse 14 || B:820 || C:140 || D:196400
R5| A:Weidgasse 16 || B:760 || D:172800
R7| A:Zwischensumme || B:1580 || C:140 || D:369200"""


def xlsx(build) -> bytes:
    wb = Workbook()
    build(wb.active)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- §6 acceptance ----------------------------------------------------------

def test_silent_output_is_exact():
    assert serialize_xlsx(OBJEKTLISTE, "silent")["text"] == EXPECTED_SILENT


def test_empty_row_is_omitted_not_renumbered():
    lines = serialize_xlsx(OBJEKTLISTE, "silent")["text"].splitlines()
    assert not any(l.startswith("R6|") for l in lines)
    assert any(l.startswith("R7|") for l in lines)


def test_empty_cell_is_absent_not_blank():
    r5 = next(l for l in serialize_xlsx(OBJEKTLISTE, "silent")["text"].splitlines() if l.startswith("R5|"))
    assert "C:" not in r5
    assert r5 == "R5| A:Weidgasse 16 || B:760 || D:172800"


def test_number_has_no_thousands_separator():
    text = serialize_xlsx(OBJEKTLISTE, "silent")["text"]
    assert "196400" in text
    assert "196'400" not in text and "196,400" not in text


def test_merges_annotated_only_at_anchor():
    text = serialize_xlsx(OBJEKTLISTE, "silent")["text"]
    assert text.count("[merged A1:D1]") == 1
    assert text.count("[merged B2:C2]") == 1
    # the covered cells are not emitted at all
    assert "C[merged" not in text


def test_formula_mode_annotates_each_formula():
    last = serialize_xlsx(OBJEKTLISTE, "formula")["text"].splitlines()[-1]
    assert last == (
        "R7| A:Zwischensumme || B:1580 {=SUM(B4:B5)} "
        "|| C:140 {=SUM(C4:C5)} || D:369200 {=SUM(D4:D5)}"
    )


def test_formula_mode_does_not_change_other_rows():
    silent = serialize_xlsx(OBJEKTLISTE, "silent")["text"].splitlines()
    formula = serialize_xlsx(OBJEKTLISTE, "formula")["text"].splitlines()
    assert silent[:-1] == formula[:-1]


# --- escaping ---------------------------------------------------------------

def test_escaping():
    def build(ws):
        ws["A1"] = "Fläche A|B"
        ws["B1"] = r"C:\Temp\x"
        ws["C1"] = "{=nicht=formel}"

    line = serialize_xlsx(xlsx(build), "silent")["text"].splitlines()[1]
    assert line == r"R1| A:Fläche A\|B || B:C:\\Temp\\x || C:\{=nicht=formel\}"


def test_value_may_contain_colon_and_splits_on_the_first_one():
    def build(ws):
        ws["A1"] = "10:30 Uhr"

    line = serialize_xlsx(xlsx(build), "silent")["text"].splitlines()[1]
    ref, _, value = line[len("R1| "):].partition(":")
    assert ref == "A" and value == "10:30 Uhr"


def test_newlines_become_slash():
    def build(ws):
        ws["A1"] = "Zeile1\nZeile2"

    assert "A:Zeile1 / Zeile2" in serialize_xlsx(xlsx(build), "silent")["text"]


# --- value formatting -------------------------------------------------------

def test_types():
    from datetime import date, datetime, time

    def build(ws):
        ws["A1"] = date(2026, 9, 2)
        ws["B1"] = datetime(2026, 9, 2, 14, 30, 0)
        ws["C1"] = time(14, 30, 0)
        ws["D1"] = True
        ws["E1"] = 0.075
        ws["F1"] = 1580.0

    line = serialize_xlsx(xlsx(build), "silent")["text"].splitlines()[1]
    # openpyxl widens a date to midnight on write; a midnight time is emitted date-only
    assert "A:2026-09-02 |" in line
    assert "B:2026-09-02T14:30:00" in line
    assert "C:14:30:00" in line
    assert "D:TRUE" in line
    assert "E:0.075" in line          # stored value, not 7.5%
    assert "F:1580" in line           # no trailing .0


def test_whitespace_only_cell_is_kept():
    def build(ws):
        ws["A1"] = "   "
        ws["B1"] = "x"

    assert serialize_xlsx(xlsx(build), "silent")["text"].splitlines()[1] == "R1| A:    || B:x"


# --- errors -----------------------------------------------------------------

def test_error_cells_are_omitted_in_silent_and_kept_in_error_mode():
    def build(ws):
        ws["A1"] = "ok"
        ws["B1"] = "#DIV/0!"     # openpyxl types this as an error cell

    assert serialize_xlsx(xlsx(build), "silent")["text"].splitlines()[1] == "R1| A:ok"
    assert "B:#DIV/0!" in serialize_xlsx(xlsx(build), "error")["text"]


# --- sheets -----------------------------------------------------------------

def test_single_sheet_still_gets_a_header():
    assert serialize_xlsx(OBJEKTLISTE, "silent")["text"].startswith("# Sheet: Q3\n")


def test_multiple_sheets_in_workbook_order_including_empty_ones():
    wb = Workbook()
    wb.active.title = "Erst"
    wb.active["A1"] = "a"
    wb.create_sheet("Leer")
    dritt = wb.create_sheet("Dritt")
    dritt["A1"] = "c"
    buf = io.BytesIO()
    wb.save(buf)

    text = serialize_xlsx(buf.getvalue(), "silent")["text"]
    assert [l for l in text.splitlines() if l.startswith("# Sheet:")] == [
        "# Sheet: Erst",
        "# Sheet: Leer",
        "# Sheet: Dritt",
    ]


# --- csv --------------------------------------------------------------------

def test_csv_synthesises_column_letters_and_keeps_row_gaps():
    raw = b"Liegenschaft,Wohnen\nWeidgasse 14,820\n\nWeidgasse 16,760\n"
    text = serialize_csv(raw, "objekte.csv")["text"]
    assert text == (
        "# Sheet: objekte\n"
        "R1| A:Liegenschaft || B:Wohnen\n"
        "R2| A:Weidgasse 14 || B:820\n"
        "R4| A:Weidgasse 16 || B:760"
    )


def test_csv_quoted_comma_is_one_value():
    assert 'A:a,b' in serialize_csv(b'"a,b",2\n', "x.csv")["text"]


def test_tsv_uses_tab():
    assert "A:a || B:b" in serialize_csv(b"a\tb\n", "x.tsv")["text"]


# --- applicability ----------------------------------------------------------

@pytest.mark.parametrize("ext", ["pdf", "docx", "pptx", "png", "xls"])
def test_non_tabular_inputs_are_not_applicable(ext):
    out = build_cells(b"irrelevant", f"datei.{ext}", ext, "silent")
    assert out["applicable"] is False
    assert out["message"].startswith("Zellenformat nicht anwendbar")


def test_applicable_inputs_report_resolved_mode():
    out = build_cells(OBJEKTLISTE, "objektliste.xlsx", "xlsx", "formula")
    assert out["applicable"] is True
    assert out["formulaMode"] == "formula"
    assert out["sheets"][0]["name"] == "Q3"
    assert out["sheets"][0]["rows"] == 6      # 6 non-empty rows out of 7
    assert out["truncated"] is False


# --- round trip (§6.2) ------------------------------------------------------

def test_round_trip_against_a_direct_read():
    """Serialize, parse back, compare to what openpyxl reports directly."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    def unescape(s):
        out, i = [], 0
        while i < len(s):
            if s[i] == "\\" and i + 1 < len(s):
                out.append(s[i + 1])
                i += 2
            else:
                out.append(s[i])
                i += 1
        return "".join(out)

    parsed = {}
    for line in serialize_xlsx(OBJEKTLISTE, "silent")["text"].splitlines():
        if line.startswith("# Sheet:"):
            continue
        head, _, rest = line.partition("| ")
        row = int(head[1:])
        for chunk in rest.split(" || "):
            # The separator is the first ':' *after* the optional [merged ...] group —
            # the annotation contains a ':' of its own.
            m = re.match(r"([A-Z]+)(?:\[merged [^\]]+\])?:(.*)$", chunk, re.S)
            parsed[(row, m.group(1))] = unescape(m.group(2))

    ws = load_workbook(io.BytesIO(OBJEKTLISTE), data_only=True)["Q3"]
    direct = {}
    for r in ws.iter_rows():
        for cell in r:
            if cell.value not in (None, ""):
                direct[(cell.row, get_column_letter(cell.column))] = str(cell.value)

    assert parsed == direct
