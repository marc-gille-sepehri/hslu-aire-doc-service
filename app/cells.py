"""Cell-addressed serialization for spreadsheets.

One line per row that has at least one non-empty cell:

    R{row}| {col}:{value} || {col}:{value} || ...

Row numbers and column letters are the sheet's own, empty cells are omitted, and
merged cells are annotated at their anchor. See docs/delta-spec-doc-convert-cells.md.

Memory (decision D1): the cell walk uses read_only=True and merged ranges come from a
separate streaming pass over the sheet XML. load_workbook(read_only=False) — the
obvious way to reach ws.merged_cells — peaks at roughly 120x the file size, which at
the 30 MiB upload limit would OOM the 1 vCPU / 4 GB instance next to Torch.
"""
import csv as _csv
import io
import zipfile
from datetime import date, datetime, time
from xml.etree import ElementTree as ET

from openpyxl.utils import get_column_letter

# Formats that can be addressed by cell. .xls is deliberately excluded (D3): xlrd has
# no streaming mode and cannot expose formula text.
CELL_EXTS = {"xlsx", "xlsm"}
CSV_EXTS = {"csv", "tsv"}

OUTPUT_FORMATS = ("markdown", "cells", "both")
FORMULA_MODES = ("silent", "error", "formula")

MAX_ROWS_PER_SHEET = 5000
MAX_SHEETS = 50
MAX_TEXT_BYTES = 4 * 1024 * 1024
# A merge covering more cells than this is annotated at its anchor but not used to
# suppress covered cells — building the lookup would cost more than the merge is worth.
MAX_COVERED_CELLS = 200_000

NOT_APPLICABLE = "Zellenformat nicht anwendbar — diese Datei hat keine Tabellenstruktur."

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_NS_PKG = "{http://schemas.openxmlformats.org/package/2006/relationships}"


def wants(output_format: str, part: str) -> bool:
    return output_format == part or output_format == "both"


def _escape(s: str) -> str:
    # Backslash first — the reverse order double-escapes an existing backslash.
    for a, b in (("\\", "\\\\"), ("|", "\\|"), ("{", "\\{"), ("}", "\\}")):
        s = s.replace(a, b)
    return s


def _fmt(v) -> str:
    """Stored value, never the displayed one: no thousands separators, no rounding."""
    if v is True:
        return "TRUE"
    if v is False:
        return "FALSE"
    if isinstance(v, datetime):
        # Excel has no date-only type — a date cell round-trips as midnight. Emitting
        # the date alone is right far more often than not; the cost is that a genuine
        # midnight timestamp loses its T00:00:00.
        return v.date().isoformat() if v.time() == time(0, 0) else v.isoformat(sep="T", timespec="seconds")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat(timespec="seconds")
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return _escape(str(v).replace("\r\n", "\n").replace("\n", " / "))


def _sheet_parts(raw: bytes) -> dict[str, str]:
    """sheet name -> path of its XML part inside the xlsx zip."""
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        book = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        target = {r.get("Id"): r.get("Target") for r in rels.findall(_NS_PKG + "Relationship")}
        parts = {}
        for sheet in book.find(_NS + "sheets") if book.find(_NS + "sheets") is not None else []:
            t = target.get(sheet.get(_NS_REL + "id")) or ""
            t = t.lstrip("/")
            if not t.startswith("xl/"):
                t = "xl/" + t
            parts[sheet.get("name")] = t
        return parts


def _merge_refs(raw: bytes, part: str) -> list[str]:
    """Merged ranges without loading the sheet — the whole point of D1.

    <mergeCells> follows <sheetData> in the schema, so this walks the row elements;
    clearing each one on close keeps the pass flat in memory."""
    refs: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z, z.open(part) as f:
            for _, el in ET.iterparse(f, events=("end",)):
                if el.tag == _NS + "mergeCell":
                    refs.append(el.get("ref"))
                if el.tag in (_NS + "row", _NS + "c", _NS + "mergeCell"):
                    el.clear()
    except KeyError:
        pass  # sheet part missing — no merges rather than a failed conversion
    return refs


def _bounds(ref: str) -> tuple[int, int, int, int]:
    from openpyxl.utils import range_boundaries

    min_c, min_r, max_c, max_r = range_boundaries(ref)
    return min_r, min_c, max_r, max_c


def _merge_index(refs: list[str]) -> tuple[dict[str, str], set[tuple[int, int]], list[str]]:
    """anchor coord -> ref, plus the set of cells a merge covers (anchor excluded)."""
    anchors: dict[str, str] = {}
    covered: set[tuple[int, int]] = set()
    warnings: list[str] = []
    total = 0
    for ref in refs:
        min_r, min_c, max_r, max_c = _bounds(ref)
        anchors[f"{get_column_letter(min_c)}{min_r}"] = ref
        total += (max_r - min_r + 1) * (max_c - min_c + 1)
        if total > MAX_COVERED_CELLS:
            warnings.append("sehr grosse verbundene Bereiche — überdeckte Zellen werden mitausgegeben")
            continue
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                covered.add((r, c))
    return anchors, covered, warnings


def _formulas(raw: bytes, sheet_name: str) -> dict[str, str]:
    """coord -> formula text. Second streaming pass; only for formulaMode 'formula'."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    try:
        ws = wb[sheet_name]
        out = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value is not None:
                    out[cell.coordinate] = str(cell.value)
        return out
    finally:
        wb.close()


def _serialize_sheet(ws, anchors, covered, formulas, mode) -> tuple[str, int, int, bool]:
    lines: list[str] = []
    n_cells = 0
    truncated = False
    for row in ws.iter_rows():
        if not row:
            continue
        if len(lines) >= MAX_ROWS_PER_SHEET:
            truncated = True
            break
        parts = []
        rownum = None
        for cell in row:
            # read_only mode yields EmptyCell for gaps — it carries no coordinate or
            # data_type, so nothing but .value may be touched before this check.
            v = cell.value
            if v is None or v == "":
                continue
            coord = cell.coordinate
            if (cell.row, cell.column) in covered and coord not in anchors:
                continue
            is_error = cell.data_type == "e"
            if is_error and mode == "silent":
                continue
            rownum = cell.row
            col = get_column_letter(cell.column)
            ref = f"{col}[merged {anchors[coord]}]" if coord in anchors else col
            text = str(v) if is_error else _fmt(v)
            if mode == "formula":
                f = formulas.get(coord)
                if f:
                    text = f"{text} {{{_escape(f)}}}"
            parts.append(f"{ref}:{text}")
            n_cells += 1
        if parts:
            lines.append(f"R{rownum}| " + " || ".join(parts))
    return "\n".join(lines), len(lines), n_cells, truncated


def serialize_xlsx(raw: bytes, formula_mode: str) -> dict:
    from openpyxl import load_workbook

    parts = _sheet_parts(raw)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets, warnings, blocks = [], [], []
    truncated = False
    total_bytes = 0
    try:
        names = wb.sheetnames
        if len(names) > MAX_SHEETS:
            warnings.append(f"nur die ersten {MAX_SHEETS} von {len(names)} Blättern")
            names = names[:MAX_SHEETS]
            truncated = True
        for name in names:
            refs = _merge_refs(raw, parts.get(name, ""))
            anchors, covered, warn = _merge_index(refs)
            warnings.extend(warn)
            formulas = _formulas(raw, name) if formula_mode == "formula" else {}
            text, n_rows, n_cells, cut = _serialize_sheet(
                wb[name], anchors, covered, formulas, formula_mode
            )
            if cut:
                truncated = True
                text += f"\n… [gekürzt: {MAX_ROWS_PER_SHEET} Zeilen]"
            total_bytes += len(text.encode("utf-8"))
            if total_bytes > MAX_TEXT_BYTES:
                warnings.append("Ausgabe an der Grössengrenze abgeschnitten")
                truncated = True
                break
            sheets.append({"name": name, "text": text, "rows": n_rows, "cells": n_cells})
            blocks.append(f"# Sheet: {name}\n{text}" if text else f"# Sheet: {name}")
    finally:
        wb.close()

    if formula_mode in ("formula", "error") and not any(s["cells"] for s in sheets):
        warnings.append("keine gespeicherten Formelergebnisse — Datei zuletzt nicht von Excel gespeichert")

    return {
        "text": "\n".join(blocks),
        "sheets": sheets,
        "formulaMode": formula_mode,
        "truncated": truncated,
        "warnings": warnings,
    }


def serialize_csv(raw: bytes, filename: str) -> dict:
    delimiter = "\t" if filename.lower().endswith(".tsv") else ","
    text = raw.decode("utf-8-sig", errors="replace")
    name = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "Tabelle"

    lines, n_cells, truncated = [], 0, False
    for i, fields in enumerate(_csv.reader(io.StringIO(text), delimiter=delimiter), start=1):
        if len(lines) >= MAX_ROWS_PER_SHEET:
            truncated = True
            break
        parts = []
        for j, raw_value in enumerate(fields, start=1):
            if raw_value == "":
                continue
            parts.append(f"{get_column_letter(j)}:{_fmt(raw_value)}")
            n_cells += 1
        if parts:
            lines.append(f"R{i}| " + " || ".join(parts))

    body = "\n".join(lines)
    if truncated:
        body += f"\n… [gekürzt: {MAX_ROWS_PER_SHEET} Zeilen]"
    return {
        "text": f"# Sheet: {name}\n{body}",
        "sheets": [{"name": name, "text": body, "rows": len(lines), "cells": n_cells}],
        "formulaMode": "silent",  # a CSV has no formulas
        "truncated": truncated,
        "warnings": [],
    }


def not_applicable() -> dict:
    return {"applicable": False, "reason": "no_table_structure", "message": NOT_APPLICABLE}


def build_cells(data: bytes, filename: str, ext: str, formula_mode: str) -> dict:
    """Dispatch on extension; never raises for an inapplicable input (§3)."""
    if ext in CELL_EXTS:
        out = serialize_xlsx(data, formula_mode)
    elif ext in CSV_EXTS:
        out = serialize_csv(data, filename)
    else:
        return not_applicable()
    out["applicable"] = True
    out["requestedFormulaMode"] = formula_mode
    return out
